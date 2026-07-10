import json
import re
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

from docx import Document
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = Path(r"C:/Users/HUAWEI/Downloads/Brandream_Instagram_Influencer_Tracker_40.xlsx")
DOCX_PATH = Path(r"C:/Users/HUAWEI/Downloads/Brandream_家居床品品牌竞品可视化报告_2026-07.docx")
OUTPUT_JSON = ROOT / "work" / "generated_data.json"
OUTPUT_JS = ROOT / "seed-data.js"

NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def to_iso_date(value):
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return ""


def derive_tier(count):
    count = count or 0
    if count < 10000:
        return "Nano"
    if count < 50000:
        return "Micro"
    if count < 150000:
        return "Mid-tier"
    if count < 500000:
        return "Macro"
    return "Mega"


def load_influencers():
    now = datetime.now().isoformat()
    wb = load_workbook(XLSX_PATH, data_only=False)
    ws = wb["达人总表"]
    headers = [cell.value for cell in ws[1]]
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        item = dict(zip(headers, row))
        handle = item.get("达人账号") or ""
        follower_count = int(item.get("粉丝数") or 0)
        is_contacted = str(item.get("是否已联系") or "").strip() in {"是", "已联系", "true", "True"}
        status = item.get("状态") or "未联系"
        profile_url = item.get("Instagram URL") or item.get("数据来源URL") or ""
        row_id = f"influencer-{int(item.get('序号') or len(rows) + 1)}"

        created_at = to_iso_date(item.get("计划联系日期")) or now
        updated_at = to_iso_date(item.get("实际联系日期")) or created_at
        rows.append(
            {
                "id": row_id,
                "platform": item.get("平台") or "Instagram",
                "accountHandle": handle,
                "displayName": item.get("达人/账号名称") or handle.replace("@", "") or "Unknown Creator",
                "profileUrl": profile_url,
                "country": item.get("地区/国家") or "",
                "followerCount": follower_count,
                "followerTier": item.get("粉丝级别") or derive_tier(follower_count),
                "contentType": item.get("内容类型") or "",
                "bioKeywords": item.get("主页简介/关键词") or "",
                "avatarUrl": "",
                "autoFillStatus": "来自 Excel 导入",
                "fitReason": item.get("匹配理由") or "",
                "recommendedProducts": item.get("推荐产品") or "",
                "contactInfo": item.get("公开邮箱/联系方式") or "",
                "priority": item.get("建议优先级") or "",
                "status": status,
                "isContacted": is_contacted,
                "firstContactChannel": item.get("初次联系渠道") or "",
                "quoteAmount": item.get("报价") or "",
                "notes": item.get("备注/待确认") or "",
                "isSelectedForReport": False,
                "isDeleted": False,
                "deletedAt": "",
                "deletionReason": "",
                "images": [],
                "logs": [
                    {
                        "id": f"{row_id}-log-import",
                        "actionType": "导入",
                        "actionNote": "从 Excel 导入达人主档案",
                        "actionDate": updated_at,
                    }
                ],
                "createdAt": created_at,
                "updatedAt": updated_at,
            }
        )

    return rows


def extract_docx_hyperlinks():
    with zipfile.ZipFile(DOCX_PATH) as archive:
        rels_root = ET.fromstring(archive.read("word/_rels/document.xml.rels"))
        rel_targets = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels_root.findall("pr:Relationship", NS)
            if rel.attrib.get("Type", "").endswith("/hyperlink")
        }
        document_root = ET.fromstring(archive.read("word/document.xml"))

    entries = []
    for hyperlink in document_root.findall(".//w:hyperlink", NS):
        rid = hyperlink.attrib.get(f"{{{NS['r']}}}id")
        text = "".join(hyperlink.itertext()).strip()
        target = rel_targets.get(rid, "")
        if target:
            entries.append({"text": text, "url": target})
    return entries


def load_brand_tables():
    doc = Document(DOCX_PATH)
    table_rows = []
    for idx, table in enumerate(doc.tables):
        rows = []
        for row in table.rows:
            values = [cell.text.strip() for cell in row.cells]
            if any(values):
                rows.append(values)
        table_rows.append({"index": idx, "rows": rows})
    return table_rows


def normalize_brand_name(name):
    return re.sub(r"\s+", " ", name.strip()).lower()


def load_brands():
    hyperlinks = extract_docx_hyperlinks()
    tables = load_brand_tables()
    brand_names = ["Ownkoti", "The Inside", "Pottery Barn", "EverGrace", "HORIMOTE HOME", "Brandream"]
    brand_map = {
        normalize_brand_name(name): {
            "id": f"brand-{re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')}",
            "brandName": name,
            "country": "",
            "category": "Bedding",
            "priceBand": "",
            "brandIntro": "",
            "amazonUrl": "",
            "websiteUrl": "",
            "parentCompany": "",
            "sourceNote": "Brandream_家居床品品牌竞品可视化报告_2026-07.docx",
            "relatedInfluencers": "",
            "notes": "",
        }
        for name in brand_names
    }

    profile_tables = {
        "ownkoti": 7,
        "the inside": 10,
        "pottery barn": 13,
        "evergrace": 16,
        "horimote home": 19,
    }

    country_defaults = {
        "ownkoti": "",
        "the inside": "United States",
        "pottery barn": "United States",
        "evergrace": "United States",
        "horimote home": "",
        "brandream": "",
    }

    for key, table_index in profile_tables.items():
        rows = tables[table_index]["rows"]
        brand = brand_map[key]
        brand["country"] = country_defaults.get(key, "")
        for row in rows:
            if len(row) < 2:
                continue
            field, value = row[0], row[1]
            if "品牌定位" in field:
                brand["brandIntro"] = value
            elif "主要渠道" in field:
                brand["notes"] = value
            elif "公司信息" in field or "历史节点" in field:
                brand["parentCompany"] = value
            elif "公开规模" in field:
                brand["priceBand"] = value

    # Manual category refinement from document positioning
    brand_map["ownkoti"]["category"] = "Home decor / Bedding"
    brand_map["the inside"]["category"] = "Designer home / Upholstery"
    brand_map["pottery barn"]["category"] = "Full-home retail / Bedding"
    brand_map["evergrace"]["category"] = "Bedding / Textile"
    brand_map["horimote home"]["category"] = "Bedding / Essentials"
    brand_map["brandream"]["category"] = "Bedding"

    for link in hyperlinks:
        text = normalize_brand_name(link["text"])
        url = link["url"]
        for key, brand in brand_map.items():
            if key in text or text in key:
                lowered_url = url.lower()
                if "amazon." in lowered_url and not brand["amazonUrl"]:
                    brand["amazonUrl"] = url
                elif not brand["websiteUrl"]:
                    brand["websiteUrl"] = url

    # Prefer direct doc hyperlinks where present.
    for link in hyperlinks:
        url = link["url"]
        if "theinside.com/about-us" in url:
            brand_map["the inside"]["websiteUrl"] = url
        elif "evergracehome.com" in url:
            brand_map["evergrace"]["websiteUrl"] = url
        elif "amazon.com/EverGrace" in url:
            brand_map["evergrace"]["amazonUrl"] = url
        elif "horimotehome.com" in url:
            brand_map["horimote home"]["websiteUrl"] = url
        elif "amazon.com/HORIMOTE-HOME" in url:
            brand_map["horimote home"]["amazonUrl"] = url
        elif "ownkoti.com/pages/about-ownkoti" in url:
            brand_map["ownkoti"]["websiteUrl"] = url
        elif "amazon.com/Brandream-Bedding" in url:
            brand_map["brandream"]["amazonUrl"] = url

    # Fallback official websites where document text implies brand pages but hyperlink text may not mention brand directly
    fallback_urls = {
        "ownkoti": {
            "websiteUrl": "https://www.ownkoti.com/pages/about-ownkoti",
            "amazonUrl": "",
        },
        "the inside": {
            "websiteUrl": "https://www.theinside.com/about-us",
            "amazonUrl": "",
        },
        "pottery barn": {
            "websiteUrl": "https://www.potterybarn.com",
            "amazonUrl": "",
        },
        "evergrace": {
            "websiteUrl": "https://evergracehome.com/",
            "amazonUrl": "https://www.amazon.com/EverGrace-Pre-Washed-Lightweight-Bedspread-Oversized/dp/B0G4VGR2T1",
        },
        "horimote home": {
            "websiteUrl": "https://horimotehome.com/",
            "amazonUrl": "https://www.amazon.com/HORIMOTE-HOME-Pre-Washed-Lightweight-Bedspreads/dp/B0GBSSRCJF",
        },
        "brandream": {
            "websiteUrl": "",
            "amazonUrl": "https://www.amazon.com/Brandream-Bedding-Woodland-Nursery-Printed/dp/B0CMJYLYB8",
        },
    }

    for key, fallback in fallback_urls.items():
        brand = brand_map[key]
        if not brand["websiteUrl"]:
            brand["websiteUrl"] = fallback["websiteUrl"]
        if not brand["amazonUrl"]:
            brand["amazonUrl"] = fallback["amazonUrl"]

    return list(brand_map.values())


def main():
    payload = {
        "version": datetime.now().isoformat(),
        "influencers": load_influencers(),
        "brands": load_brands(),
    }
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    OUTPUT_JS.write_text(
        "window.__APP_IMPORTED_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(str(OUTPUT_JSON))


if __name__ == "__main__":
    main()
